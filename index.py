import os
import pandas as pd
import requests
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Function to check and convert .xls to .xlsx
def check_and_convert_to_xlsx(file_path):
    if file_path.endswith('.xls'):
        print(f"Converting {file_path} to .xlsx format...")
        df = pd.read_excel(file_path, skiprows=4)  # Read the .xls file
        new_file_path = file_path.replace('.xls', '.xlsx')  # Create new file path for .xlsx
        df.to_excel(new_file_path, index=False)  # Save it as .xlsx
        print(f"File converted to {new_file_path}")
        return new_file_path  # Return the new .xlsx file path
    return file_path  # Return original path if it's already .xlsx

# Path to your Excel file
file_path = r'./NOTAMsExcel.xls'  # Replace with your actual file path if needed

# Check and convert .xls to .xlsx if needed
file_path = check_and_convert_to_xlsx(file_path)

# Load the Excel file (now in .xlsx format)
df = pd.read_excel(file_path, skiprows=4)

# Extract relevant columns
notam_details = df.iloc[:, [0, 4, 5, 6]]
notam_details.columns = ['Airport ID', 'Effective Date', 'Expiration Date', 'NOTAM Line']

# ===========================
# METAR API Connection Section
# ===========================

# Function to fetch METAR data for a given airport ID
def get_metar_data(airport_id, format):
    base_url = "https://aviationweather.gov/api/data/metar"
    params = {
        'ids': airport_id,
        'format': format
    }
    try:
        response = requests.get(base_url, params=params)
        response.raise_for_status()  # Raise an error for bad responses
        return response.json()  # Return the JSON response
    except requests.exceptions.RequestException as e:
        print(f"Error fetching data for {airport_id}: {e}")
        return None

# Function to save JSON data to a file (Used to verify info while testing)
def save_json_data(data, airport_id):
    folder_name = 'METAR-json'
    os.makedirs(folder_name, exist_ok=True)  # Create the folder if it doesn't exist
    file_path = os.path.join(folder_name, f"{airport_id}.json")  # File name based on airport ID

    with open(file_path, 'w') as json_file:
        json.dump(data, json_file, indent=4)  # Save with pretty formatting
    print(f"Data saved to {file_path}")

# Check if a NOTAM is unserviceable, wind-related, or ceiling-related
def flag_notam(notam_line, airport_id):
    # Look for the substrings "SVC AUTOMATED WX BCST SYSTEM U/S", "SYSTEM NOT AVBL", "WIND U/S", and "CEILING UNREL"
    if "SYSTEM U/S" in notam_line or "SYSTEM NOT AVBL" in notam_line:
        #print(f"Unserviceable NOTAM found for {airport_id}. Checking METAR data...")
        metar_data = get_metar_data("K"+airport_id, "json")
        
        # If the METAR API returns no data, the flag is OK
        if metar_data:
            #print(f"METAR data found for {airport_id}. Flagging NOTAM.\n")
            return "FLAGGED"
        else:
            #print(f"No METAR data for {airport_id}. NOTAM is valid as unserviceable.\n")
            return "OK"

    elif "WIND" in notam_line:
        #print(f"WIND U/S found for {airport_id}. Checking METAR wind data...")
        metar_data = get_metar_data("K"+airport_id, "json")

        if metar_data:
            wind_data = metar_data[0]

            #If we should only check for Wind Gust, adjust params as such, else continue for all wind params
            if "GUST" in notam_line:
                wgst = wind_data.get('wgst', None)
                if wgst is None:
                    return "OK"
                else:
                    return "FLAGGED"

            wdir = wind_data.get('wdir', None)
            wspd = wind_data.get('wspd', None)
            wgst = wind_data.get('wgst', None)

            if wdir is None and wspd is None and wgst is None:
                #print(f"WIND U/S confirmed for {airport_id}. No flag needed.\n")
                return "OK"
            else:
                #print(f"METAR data for wind exists at {airport_id}. Flagging NOTAM.\n")
                return "FLAGGED"
        else:
            #print(f"No METAR data for {airport_id}. WIND U/S status accepted.\n")
            return "OK"
    
    elif "CEILING" in notam_line:
        #print(f"CEILING UNREL found for {airport_id}. Checking METAR ceiling data...")
        metar_data = get_metar_data("K"+airport_id, "json")

        if metar_data:
            # Check if clouds object is empty
            clouds = metar_data[0].get('clouds', [])
            if not clouds:  # If clouds is an empty list
                #print(f"CEILING UNREL confirmed for {airport_id}. No flag needed.\n")
                return "OK"
            else:
                #print(f"Clouds data exists at {airport_id}. Flagging NOTAM.\n")
                return "FLAGGED"
        else:
            #print(f"No METAR data for {airport_id}. CEILING UNREL status accepted.\n")
            return "OK"

        # Check for "WX UNREL"
    elif "PRESENT WX" in notam_line:
        #print(f"WX UNREL found for {airport_id}. Checking METAR data...")
        metar_data = get_metar_data("K" + airport_id, "json")

        if metar_data:
            wx_data = metar_data[0]
            temp = wx_data.get('temp', None)
            dewp = wx_data.get('dewp', None)
            Visib = wx_data.get('visib', None)
            wxstring = wx_data.get('wxString', None)

            # If temp, dewp, and vertVis are not null, flag the NOTAM
            if temp is not None or dewp is not None or Visib is not None or wxstring is not None:
                #print(f"Weather data found for {airport_id}. Flagging NOTAM.\n")
                return "FLAGGED"
            else:
                #print(f"Weather data is unreliable for {airport_id}. No flag needed.\n")
                return "OK"
        else:
            #print(f"No METAR data for {airport_id}. WX UNREL status accepted.\n")
            return "OK"

    # Check for "ALTIMETER SETTING UNREL"
    elif "ALTIMETER SETTING" in notam_line:
        #print(f"ALTIMETER SETTING UNREL found for {airport_id}. Checking METAR altimeter data...")
        metar_data = get_metar_data("K" + airport_id, "json")

        if metar_data:
            altimeter_data = metar_data[0]
            altim = altimeter_data.get('altim', None)

            # If altim is not null, flag the NOTAM
            if altim is not None:
                #print(f"Altimeter data found for {airport_id}. Flagging NOTAM.\n")
                return "FLAGGED"
            else:
                #print(f"Altimeter data is unreliable for {airport_id}. No flag needed.\n")
                return "OK"
        else:
            #print(f"No METAR data for {airport_id}. ALTIMETER SETTING UNREL status accepted.\n")
            return "OK"
        
    # Check for "PRECIPITATION"
    elif "PRECIPITATION" in notam_line:
        #print(f"PRECIPITATION found for {airport_id}. Checking METAR precipitation data...")
        metar_data = get_metar_data("K" + airport_id, "json")

        if metar_data:
            precip_data = metar_data[0]
            precip = precip_data.get('precip', None)

            # If precip is not null, flag the NOTAM
            if precip is not None:
                #print(f"Precipitation data found for {airport_id}. Flagging NOTAM.\n")
                return "FLAGGED"
            else:
                #print(f"No precipitation data for {airport_id}. No flag needed.\n")
                return "OK"
        else:
            #print(f"No METAR data for {airport_id}. PRECIPITATION status accepted.\n")
            return "OK"

    # Check for "VIS"
    elif "VIS" in notam_line:
        #print(f"VIS found for {airport_id}. Checking METAR visibility data...")
        metar_data = get_metar_data("K" + airport_id, "json")

        if metar_data:
            visibility_data = metar_data[0]
            visib = visibility_data.get('visib', None)

            # If visib is not null, flag the NOTAM
            if visib is not None:
                #print(f"Visibility data found for {airport_id}. Flagging NOTAM.\n")
                return "FLAGGED"
            else:
                #print(f"No visibility data for {airport_id}. No flag needed.\n")
                return "OK"
        else:
            #print(f"No METAR data for {airport_id}. VIS status accepted.\n")
            return "OK"
    
    return None

# ===========================
# Excel Highlighting Section
# ===========================

# Load the workbook and select the first sheet
def highlight_flagged_rows(file_path, flagged_indices):
    wb = load_workbook(file_path)
    ws = wb.active  # Assuming the first sheet

    # Define a red fill for highlighting
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Iterate over the flagged indices and apply red highlight
    for row in flagged_indices:
        ws[f'A{row+6}'].fill = red_fill  # Adjust by 6 due to the header rows in the original Excel

    # Save the workbook with highlights
    wb.save(file_path)
    print(f"Flagged rows highlighted in {file_path}")

# Iterate through the NOTAMs and check for unserviceable systems
flagged_rows = []  # List to store rows that are flagged
print("Checking Data...\n")
for index, row in notam_details.iterrows():
    airport_id = row['Airport ID']
    notam_line = row['NOTAM Line']
    
    # Check the NOTAM line for unserviceable status and flag if needed
    
    status = flag_notam(notam_line, airport_id)
    if status == "FLAGGED":
        flagged_rows.append(index)  # Add row to the flagged list

# Highlight flagged rows in the Excel file
print("Highlighting Rows...\n")
if flagged_rows:
    highlight_flagged_rows(file_path, flagged_rows)

# Summary of flagged rows
print(f"Total flagged NOTAMs: {len(flagged_rows)}")
